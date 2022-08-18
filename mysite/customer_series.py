from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.http import HttpResponse
from datetime import datetime

def cseries(data_file):
    cseribook = Workbook()
    cserisheet = cseribook.active

    cserisheet["A1"] = "SeriesName"
    cserisheet["B1"] = "SKU"
    cserisheet["C1"] = "CustomerFirstName"
    cserisheet["D1"] = "CustomerLastName"
    cserisheet["E1"] = "CustomerGUID"
    cserisheet["F1"] = "OriginalQuantity"
    cserisheet["G1"] = "QuantityUsed"
    cserisheet["H1"] = "PurchasePrice"
    cserisheet["I1"] = "DateIssued"
    cserisheet["J1"] = "ExpirationDate"
    cserisheet["K1"] = "SeriesNumber"

    copyseribook = load_workbook(data_file)
    copyserisheet = copyseribook.active
    copyserisheet.delete_rows(idx=1, amount=2)

    allrows = copyserisheet.max_row

    for idx in range(1, allrows):
        series_name = copyserisheet.cell(row=idx + 1, column=1)
        cserisheet.cell(row=idx + 1, column=1).value = "b" + series_name.value

        sku = copyserisheet.cell(row=idx + 1, column=2)
        cserisheet.cell(row=idx + 1, column=2).value = "b" + sku.value

        first_name = copyserisheet.cell(row=idx + 1, column=3)
        cserisheet.cell(row=idx + 1, column=3).value = first_name.value

        last_name = copyserisheet.cell(row=idx + 1, column=4)
        cserisheet.cell(row=idx + 1, column=4).value = last_name.value

        guid = copyserisheet.cell(row=idx + 1, column=5)
        cserisheet.cell(row=idx + 1, column=5).value = guid.value

        original_quantity = copyserisheet.cell(row=idx + 1, column=6)
        cserisheet.cell(row=idx + 1, column=6).value = original_quantity.value

        used_quantity = copyserisheet.cell(row=idx + 1, column=7)
        cserisheet.cell(row=idx + 1, column=7).value = used_quantity.value

        price = copyserisheet.cell(row=idx + 1, column=8)
        cserisheet.cell(row=idx + 1, column=8).value = price.value

        issued = copyserisheet.cell(row=idx + 1, column=9)
        # Converts issued date to shortformat
        issuedf = f"{issued.value.month}/{issued.value.day}/{issued.value.year}"
        cserisheet.cell(row=idx + 1, column=9).value = issuedf

        expiration = copyserisheet.cell(row=idx + 1, column=10)
        expirationf = f"{expiration.value.month}/{expiration.value.day}/{expiration.value.year}"
        cserisheet.cell(row=idx + 1, column=10).value = expirationf

        series_number = copyserisheet.cell(row=idx + 1, column=11)
        cserisheet.cell(row=idx + 1, column=11).value = series_number.value


    response = HttpResponse(content=save_virtual_workbook(cseribook), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=customer-series-output.xlsx'
    return response