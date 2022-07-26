from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.http import HttpResponse

def products(data_file):
    newprodbook = Workbook()
    newprodsheet = newprodbook.active

    newprodsheet["A1"] = "Name"
    newprodsheet["B1"] = "Category"
    newprodsheet["C1"] = "SubCategory"
    newprodsheet["D1"] = "Brand"
    newprodsheet["E1"] = "Description"
    newprodsheet["F1"] = "Ingredients"
    newprodsheet["G1"] = "SKU"
    newprodsheet["H1"] = "Barcode"
    newprodsheet["I1"] = "Size/Vol"
    newprodsheet["J1"] = "UnitOfMeasure"
    newprodsheet["K1"] = "Color"
    newprodsheet["L1"] = "BuyPrice"
    newprodsheet["M1"] = "SellPrice"
    newprodsheet["N1"] = "SellOnline"
    newprodsheet["O1"] = "Professional"
    newprodsheet["P1"] = "AddOn"

    prodbook = load_workbook(data_file)
    prodsheet = prodbook.active
    prodsheet.delete_cols(idx=15, amount=9)
    prodsheet.delete_cols(idx=5, amount=1)

    mr = prodsheet.max_row
    mc = prodsheet.max_column

    for rowz in range(1, mr + 1):
        for colz in range(1, mc + 1):
            tocopy = prodsheet.cell(row=rowz + 2, column=colz)
            newprodsheet.cell(row=rowz + 1, column=colz).value = tocopy.value

    response = HttpResponse(content=save_virtual_workbook(newprodbook), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=products-output.xlsx'
    return response


def inventory(data_file):
    inventory = load_workbook(data_file)
    inventory = inventory.active

    invoutput = Workbook()
    invoutsheet = invoutput.active
    invrows = inventory.max_row

    for idx in range(1, invrows + 1):
        barcode = inventory.cell(row=idx + 2, column=9)
        invoutsheet.cell(row=idx, column=1).value = barcode.value

        stock = inventory.cell(row=idx + 2, column=15).value

        if stock == None or int(stock) <= 0:
            invoutsheet.cell(row=idx, column=2).value = 0
        elif stock == "Unlimited":
            invoutsheet.cell(row=idx, column=2).value = 999
        elif int(stock) > 0:
            invoutsheet.cell(row=idx, column=2).value = stock


    response = HttpResponse(content=save_virtual_workbook(invoutput), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=inventory-output.xlsx'
    return response

