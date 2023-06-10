from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.http import HttpResponse
#import csv

# EMPLOYEES

# CREATE NEW EMPLOYEES WORKBOOK

def employees(data_file):

    empbook = Workbook()
    toempsheet = empbook.active

    toempsheet["A1"] = "FirstName"
    toempsheet["B1"] = "LastName"
    toempsheet["C1"] = "DisplayName"
    toempsheet["D1"] = "Gender"
    toempsheet["E1"] = "Email"
    toempsheet["F1"] = "Type"
    toempsheet["G1"] = "Status"
    toempsheet["H1"] = "LicenseNumber"
    toempsheet["I1"] = "LicenseExpirationDate"
    toempsheet["J1"] = "ServiceTypeID"
    toempsheet["K1"] = "Street1"
    toempsheet["L1"] = "Street2"
    toempsheet["M1"] = "City"
    toempsheet["N1"] = "State"
    toempsheet["O1"] = "PostalCode"
    toempsheet["P1"] = "Country"
    toempsheet["Q1"] = "HomePhoneFormat"
    toempsheet["R1"] = "HomePhone"
    toempsheet["S1"] = "MobilePhoneFormat"
    toempsheet["T1"] = "MobilePhone"
    toempsheet["U1"] = "PreferredPhone"
    toempsheet["V1"] = "EmploymentEndDate"
    toempsheet["W1"] = "Notes"
	# DAY 1
    toempsheet["X1"] = "WeekDayID1"
    toempsheet["Y1"] = "DayScheduleType1"
    toempsheet["Z1"] = "StartTime1"
    toempsheet["AA1"] = "EndTime1"
    toempsheet["AB1"] = "LunchStartTime1"
    toempsheet["AC1"] = "LunchEndTime1"
	# DAY 2
    toempsheet["AD1"] = "WeekDayID2"
    toempsheet["AE1"] = "DayScheduleType2"
    toempsheet["AF1"] = "StartTime2"
    toempsheet["AG1"] = "EndTime2"
    toempsheet["AH1"] = "LunchStartTime2"
    toempsheet["AI1"] = "LunchEndTime2"
	# DAY 3
    toempsheet["AJ1"] = "WeekDayID3"
    toempsheet["AK1"] = "DayScheduleType3"
    toempsheet["AL1"] = "StartTime3"
    toempsheet["AM1"] = "EndTime3"
    toempsheet["AN1"] = "LunchStartTime3"
    toempsheet["AO1"] = "LunchEndTime3"
	# DAY 4
    toempsheet["AP1"] = "WeekDayID4"
    toempsheet["AQ1"] = "DayScheduleType4"
    toempsheet["AR1"] = "StartTime4"
    toempsheet["AS1"] = "EndTime4"
    toempsheet["AT1"] = "LunchStartTime4"
    toempsheet["AU1"] = "LunchEndTime4"
	# DAY 5
    toempsheet["AV1"] = "WeekDayID5"
    toempsheet["AW1"] = "DayScheduleType5"
    toempsheet["AX1"] = "StartTime5"
    toempsheet["AY1"] = "EndTime5"
    toempsheet["AZ1"] = "LunchStartTime5"
    toempsheet["BA1"] = "LunchEndTime5"
	# DAY 6
    toempsheet["BB1"] = "WeekDayID6"
    toempsheet["BC1"] = "DayScheduleType6"
    toempsheet["BD1"] = "StartTime6"
    toempsheet["BE1"] = "EndTime6"
    toempsheet["BF1"] = "LunchStartTime6"
    toempsheet["BG1"] = "LunchEndTime6"
	# DAY 7
    toempsheet["BH1"] = "WeekDayID7"
    toempsheet["BI1"] = "DayScheduleType7"
    toempsheet["BJ1"] = "StartTime7"
    toempsheet["BK1"] = "EndTime7"
    toempsheet["BL1"] = "LunchStartTime7"
    toempsheet["BM1"] = "LunchEndTime7"

    toempsheet["BN1"] = "EmployeeGroup"
    toempsheet["BO1"] = "AppointmentInterval"

    # OPEN AND CLEAN WORKBOOK TO COPY FROM

    copyempbook = load_workbook(data_file)
    fromempsheet = copyempbook.active
    fromempsheet.delete_rows(idx=1, amount=3)

    # COPY CELLS FROM PREVIOUS WORKBOOK
    allrows = fromempsheet.max_row

    for idx in range(1, allrows + 1):
        first_name = fromempsheet.cell(row=idx, column=1)
        toempsheet.cell(row=idx + 1, column=1).value = first_name.value

        last_name = fromempsheet.cell(row=idx, column=2)
        toempsheet.cell(row=idx + 1, column=2).value = last_name.value

		# display name does not get exported
		# column 3 stays blank

        gender = fromempsheet.cell(row=idx, column=3)
        toempsheet.cell(row=idx + 1, column=4).value = gender.value

        email = fromempsheet.cell(row=idx, column=4)
        toempsheet.cell(row=idx + 1, column=5).value = email.value

        service_type = fromempsheet.cell(row=idx, column=5)
        toempsheet.cell(row=idx + 1, column=6).value = service_type.value

        status = fromempsheet.cell(row=idx, column=6)
        toempsheet.cell(row=idx + 1, column=7).value = status.value

        license_num = fromempsheet.cell(row=idx, column=7)
        toempsheet.cell(row=idx + 1, column=8).value = license_num.value

        license_exp = fromempsheet.cell(row=idx, column=8)
        toempsheet.cell(row=idx + 1, column=9).value = license_exp.value

        type_id = fromempsheet.cell(row=idx, column=9)
        toempsheet.cell(row=idx + 1, column=10).value = type_id.value

        street1 = fromempsheet.cell(row=idx, column=10)
        toempsheet.cell(row=idx + 1, column=11).value = street1.value

        street2 = fromempsheet.cell(row=idx, column=11)
        toempsheet.cell(row=idx + 1, column=12).value = street2.value

        city = fromempsheet.cell(row=idx, column=12)
        toempsheet.cell(row=idx + 1, column=13).value = city.value

        state = fromempsheet.cell(row=idx, column=13)
        toempsheet.cell(row=idx + 1, column=14).value = state.value

        postal = fromempsheet.cell(row=idx, column=14)
        toempsheet.cell(row=idx + 1, column=15).value = postal.value

        country = fromempsheet.cell(row=idx, column=15)
        toempsheet.cell(row=idx + 1, column=16).value = country.value

        home_num1 = fromempsheet.cell(row=idx, column=16)
        toempsheet.cell(row=idx + 1, column=17).value = home_num1.value

        home_num2 = fromempsheet.cell(row=idx, column=17)
        toempsheet.cell(row=idx + 1, column=18).value = home_num2.value

        mobile_num1 = fromempsheet.cell(row=idx, column=18)
        toempsheet.cell(row=idx + 1, column=19).value = mobile_num1.value

        mobile_num2 = fromempsheet.cell(row=idx, column=19)
        toempsheet.cell(row=idx + 1, column=20).value = mobile_num2.value

        preferred_phone = fromempsheet.cell(row=idx, column=20)
        toempsheet.cell(row=idx + 1, column=21).value = preferred_phone.value

        employment_end = fromempsheet.cell(row=idx, column=21)
        toempsheet.cell(row=idx + 1, column=22).value = employment_end.value

        notes = fromempsheet.cell(row=idx, column=22)
        toempsheet.cell(row=idx + 1, column=23).value = notes.value

        toempsheet.cell(row=idx + 1, column=24).value = 1
        toempsheet.cell(row=idx + 1, column=25).value = "Day Off"

        toempsheet.cell(row=idx + 1, column=30).value = 2
        toempsheet.cell(row=idx + 1, column=31).value = "Day Off"

        toempsheet.cell(row=idx + 1, column=36).value = 3
        toempsheet.cell(row=idx + 1, column=37).value = "Day Off"

        toempsheet.cell(row=idx + 1, column=42).value = 4
        toempsheet.cell(row=idx + 1, column=43).value = "Day Off"

        toempsheet.cell(row=idx + 1, column=48).value = 5
        toempsheet.cell(row=idx + 1, column=49).value = "Day Off"

        toempsheet.cell(row=idx + 1, column=54).value = 6
        toempsheet.cell(row=idx + 1, column=55).value = "Day Off"

        toempsheet.cell(row=idx + 1, column=60).value = 7
        toempsheet.cell(row=idx + 1, column=61).value = "Day Off"

    #converts employee sheet to csv file
    #csv_version = csv.writer(open("employees-upload.csv",'w', newline = ""))
    #for row in toempsheet.rows:
    #    new_row = [cell.value for cell in row]
    #    csv_version.writerow(new_row)

# SAVE WORKBOOK
    response = HttpResponse(content=save_virtual_workbook(empbook), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=employees-output.xlsx'
    return response