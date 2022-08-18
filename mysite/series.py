from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.http import HttpResponse

def series(data_file):
    seribook = Workbook()
    serisheet = seribook.active

    serisheet["A1"] = "Name"
    serisheet["B1"] = "SKU"
    serisheet["C1"] = "Description"
    serisheet["D1"] = "Staff Fee"
    serisheet["E1"] = "Staff Fee Value"
    serisheet["F1"] = "Sell Online"
    serisheet["G1"] = "Quantity"
    serisheet["H1"] = "Unit Price"
    serisheet["I1"] = "Discount"
    serisheet["J1"] = "Sell Price"
    serisheet["K1"] = "Expires"
    serisheet["L1"] = "ExpValue"
    serisheet["M1"] = "Services"

    copyseribook = load_workbook(data_file)
    copyserisheet = copyseribook.active
    copyserisheet.delete_rows(idx=1, amount=2)

    allrows = copyserisheet.max_row

    for idx in range(1, allrows):
        a = copyserisheet.cell(row=idx + 1, column=1)
        serisheet.cell(row=idx + 1, column=1).value = "b" + a.value

        b = copyserisheet.cell(row=idx + 1, column=2)
        serisheet.cell(row=idx + 1, column=2).value = "b" + b.value

        c = copyserisheet.cell(row=idx + 1, column=3)
        serisheet.cell(row=idx + 1, column=3).value = c.value

        d = copyserisheet.cell(row=idx + 1, column=4)
        serisheet.cell(row=idx + 1, column=4).value = d.value

        e = copyserisheet.cell(row=idx + 1, column=5)
        serisheet.cell(row=idx + 1, column=5).value = e.value

        f = copyserisheet.cell(row=idx + 1, column=6)
        if f.value == 0:
            serisheet.cell(row=idx + 1, column=6).value = "No"
        elif f.value == 1:
            serisheet.cell(row=idx + 1, column=6).value = "Yes"
        else:
            serisheet.cell(row=idx + 1, column=6).value = ""

        g = copyserisheet.cell(row=idx + 1, column=7)
        serisheet.cell(row=idx + 1, column=7).value = g.value

        h = copyserisheet.cell(row=idx + 1, column=8)
        serisheet.cell(row=idx + 1, column=8).value = h.value

        i = copyserisheet.cell(row=idx + 1, column=9)
        serisheet.cell(row=idx + 1, column=9).value = i.value

        j = copyserisheet.cell(row=idx + 1, column=10)
        serisheet.cell(row=idx + 1, column=10).value = j.value

        k = copyserisheet.cell(row=idx + 1, column=11)
        serisheet.cell(row=idx + 1, column=11).value = k.value

        l = copyserisheet.cell(row=idx + 1, column=12)
        serisheet.cell(row=idx + 1, column=12).value = l.value

        m = copyserisheet.cell(row=idx + 1, column=13)
        serisheet.cell(row=idx + 1, column=13).value = m.value

    response = HttpResponse(content=save_virtual_workbook(seribook), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=series-output.xlsx'
    return response