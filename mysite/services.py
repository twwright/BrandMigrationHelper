from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.http import HttpResponse
#import csv

# SERVICES

# CREATE NEW SERVICES WORKBOOK

def services(data_file):

    servbook = Workbook()
    servsheet = servbook.active

    servsheet["A1"] = "Name"
    servsheet["B1"] = "Category"
    servsheet["C1"] = "SubCategory"
    servsheet["D1"] = "SKU"
    servsheet["E1"] = "Barcode"
    servsheet["F1"] = "Duration"
    servsheet["G1"] = "RecoveryTime"
    servsheet["H1"] = "ServiceType"
    servsheet["I1"] = "AllowCustomersToBookOnline"
    servsheet["J1"] = "Description"
    servsheet["K1"] = "RequiresTwoStaffMembers"
    servsheet["L1"] = "Price"
    servsheet["M1"] = "Addon"
    servsheet["N1"] = "SpecialEvent"

    # OPEN AND CLEAN WORKBOOK TO COPY FROM

    copyservbook = load_workbook(data_file)
    copyservsheet = copyservbook.active
    copyservsheet.delete_rows(idx=1, amount=2)

    # COPY CELLS FROM PREVIOUS WORKBOOK
    allrows = copyservsheet.max_row

    for idx in range(1, allrows + 1):
        name = copyservsheet.cell(row=idx, column=1)
        servsheet.cell(row=idx + 1, column=1).value = name.value

        category = copyservsheet.cell(row=idx, column=2)
        servsheet.cell(row=idx + 1, column=2).value = category.value

        sub_category = copyservsheet.cell(row=idx, column=3)
        servsheet.cell(row=idx + 1, column=3).value = sub_category.value

        sku = copyservsheet.cell(row=idx, column=4)
        servsheet.cell(row=idx + 1, column=10).value = sku.value

        # barcode stays empty

        duration = copyservsheet.cell(row=idx, column=9)
        servsheet.cell(row=idx + 1, column=6).value = duration.value

        recovery_time = copyservsheet.cell(row=idx, column=11)
        if recovery_time.value == None:
           servsheet.cell(row=idx + 1, column=7).value = 0
        else:
          servsheet.cell(row=idx + 1, column=7).value = recovery_time.value

        # service type stays empty

        allow_online = copyservsheet.cell(row=idx, column=7)
        if allow_online.value == 1:
           servsheet.cell(row=idx + 1, column=9).value = allow_online.value
        else:
          servsheet.cell(row=idx + 1, column=9).value = 0

        description = copyservsheet.cell(row=idx, column=4)
        servsheet.cell(row=idx + 1, column=10).value = description.value

        requires_two = copyservsheet.cell(row=idx, column=8)
        if requires_two.value == 1:
            servsheet.cell(row=idx + 1, column=11).value = requires_two.value
        else:
           servsheet.cell(row=idx + 1, column=11).value = 0

        price = copyservsheet.cell(row=idx, column=12)
        servsheet.cell(row=idx + 1, column=12).value = price.value

        # sets "Addon" value in column M
        if category.value == "Add-ons":
           servsheet.cell(row=idx + 1, column=13).value = 1
        else:
           servsheet.cell(row=idx + 1, column=13).value = 0

        # gives special_event a value in column N
        servsheet.cell(row=idx + 1, column=14).value = 0

    servsheet.delete_rows(idx=allrows, amount=2)

    #converts service sheet to csv file
    #csv_version = csv.writer(open("services-upload.csv",'w', newline = ""))
    #for row in servsheet.rows:
    #    new_row = [cell.value for cell in row]
    #    csv_version.writerow(new_row)

# SAVE WORKBOOK
    response = HttpResponse(content=save_virtual_workbook(servbook), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=services-output.xlsx'
    return response